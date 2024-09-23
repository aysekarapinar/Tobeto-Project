from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
from constant import globalConstant
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl

class Test_Tobeto_Report:
    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstant.URL)
        sleep(5)
    
    def teardown_method(self):
        self.driver.quit()


    # def getData():
    #     excelFile = openpyxl.load_workbook(globalConstant.excelLogin)
    #     sheet = excelFile["Sayfa1"] 
    #     rows = sheet.max_row
    #     data = []
    #     for i in range(2,rows+1):
    #         email = sheet.cell(i,1).value 
    #         password = sheet.cell(i,2).value
    #         data.append((email,password))
    #     return data
    
    
    def login(self, email, password):
        email_input = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, globalConstant.emailBox))
        )
        email_input.send_keys(email)

        password_input = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, globalConstant.passwordBox))
        )
        password_input.send_keys(password)

        login_button = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, globalConstant.loginButton))
        )
        login_button.click()
    
    #@pytest.mark.parametrize("email,password",getData())
    def test_S11_C1_C2_C3(self):
        self.login(globalConstant.email, globalConstant.password)
        sleep(10)
        startButton=self.driver.find_element(By.XPATH,globalConstant.startButton11XPATH)
        actions = ActionChains(self.driver)
        actions.move_to_element(startButton).perform()
        sleep(7)
        startButton.click()
        sleep(15)
        reportButton=self.driver.find_element(By.XPATH,globalConstant.raporXPATH)
        reportButton.click()
        sleep(5)

    def test_S12_C1(self):
        self.login(globalConstant.email, globalConstant.password)
        sleep(10)
        startButton2=self.driver.find_element(By.XPATH,globalConstant.startButton12XPATH)
        actions = ActionChains(self.driver)
        actions.move_to_element(startButton2).perform()
        sleep(7)
        startButton2.click()
        sleep(5)
        experience=self.driver.find_element(By.XPATH,globalConstant.experienceXPATH)
        experience.click()
        sleep(5)
        education=self.driver.find_element(By.XPATH,globalConstant.educationsXPATH)
        education.click()
        sleep(3)
        compentencies=self.driver.find_element(By.XPATH,globalConstant.compentenciesXPATH)
        compentencies.click()
        sleep(3)
        certificate=self.driver.find_element(By.XPATH,globalConstant.certificateXPATH)
        certificate.click()
        sleep(3)
        member=self.driver.find_element(By.XPATH,globalConstant.memberXPATH)
        member.click()
        sleep(3)
        prize=self.driver.find_element(By.XPATH,globalConstant.priceXPATH)
        prize.click()
        sleep(3)
        personalAccount=self.driver.find_element(By.XPATH,globalConstant.personalAccountXPATH)
        personalAccount.click()
        sleep(3)
        language=self.driver.find_element(By.XPATH,globalConstant.languageXPATH)
        language.click()
        sleep(3)
        settings=self.driver.find_element(By.XPATH,globalConstant.settingsXPATH)
        settings.click()
        sleep(2) 

    def test_S12_C1_2(self):
        self.login(globalConstant.email, globalConstant.password)
        sleep(10)
    
    elements_to_click = [
        globalConstant.startButton12XPATH,
        globalConstant.experienceXPATH,
        globalConstant.educationsXPATH,
        globalConstant.compentenciesXPATH,
        globalConstant.certificateXPATH,
        globalConstant.memberXPATH,
        globalConstant.priceXPATH,
        globalConstant.personalAccountXPATH,
        globalConstant.languageXPATH,
        globalConstant.settingsXPATH
    ]
    
    actions = ActionChains(globalConstant.startButton12XPATH)
    
    for xpath in elements_to_click:
        element = globalConstant.startButton12XPATH(By.XPATH, xpath)
        actions.move_to_element(element).perform()
        sleep(3)
        element.click()
        sleep(3)
